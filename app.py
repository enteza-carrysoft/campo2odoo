import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re
import io

st.set_page_config(page_title="Convertidor a Plantilla Campo2Odoo", layout="wide")
st.title("Convertidor a Plantilla Campo2Odoo")
st.markdown("Convierte archivos de **gastos** o **ingresos** al formato de plantilla requerido por Campo2Odoo.")

# ---------------------------------------------------------------------------
# Utilidades compartidas
# ---------------------------------------------------------------------------
def limpiar_texto(valor):
    if pd.isna(valor):
        return ""
    return str(valor).strip()


def limpiar_nif(nif):
    nif = limpiar_texto(nif)
    nif = re.sub(r"[\s\-.]", "", nif).upper()
    return nif


def formatear_fecha(valor):
    if pd.isna(valor) or str(valor).strip() == "":
        return ""
    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y")
    try:
        dt = pd.to_datetime(valor)
        return dt.strftime("%d/%m/%Y")
    except Exception:
        return str(valor).split()[0]


def calcular_vencimiento(fecha_str, dias=30):
    if not fecha_str:
        return ""
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(fecha_str, fmt)
            venc = dt + timedelta(days=dias)
            return venc.strftime("%d/%m/%Y")
        except ValueError:
            continue
    return ""


def es_numero_valido(valor):
    try:
        float(str(valor).replace(",", ".").replace(" ", ""))
        return True
    except (ValueError, TypeError):
        return False


def parsear_precio(valor):
    try:
        return float(str(valor).replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return 0.0


def parsear_iva(valor):
    iva_str = limpiar_texto(valor)
    if iva_str.upper() == "NS":
        return 0
    try:
        return int(float(iva_str))
    except (ValueError, TypeError):
        return 0


# ---------------------------------------------------------------------------
# Exportador a Excel con formato de plantilla
# ---------------------------------------------------------------------------
def exportar_excel(df_plantilla):
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas"
    ws.sheet_view.showGridLines = True

    headers = df_plantilla.columns.tolist()
    ws.append(headers)

    for _, row in df_plantilla.iterrows():
        ws.append(row.tolist())

    header_fill = PatternFill(
        start_color="1E3A5F", end_color="1E3A5F", fill_type="solid"
    )
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    col_widths = {
        "A": 28,
        "B": 20,
        "C": 18,
        "D": 16,
        "E": 20,
        "F": 40,
        "G": 12,
        "H": 16,
        "I": 18,
        "J": 12,
        "K": 26,
        "L": 14,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
            if cell.column_letter == "G":
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="center")
            elif cell.column_letter == "H":
                cell.number_format = "#,##0.00"
            elif cell.column_letter == "J":
                cell.number_format = "0"
                cell.alignment = Alignment(horizontal="center")
            elif cell.column_letter == "L":
                cell.number_format = "0"
                cell.alignment = Alignment(horizontal="center")
            elif cell.column_letter in ("B", "C", "D", "E", "I", "K"):
                cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(output)
    output.seek(0)
    return output


# ---------------------------------------------------------------------------
# Procesadores por tipo
# ---------------------------------------------------------------------------
def procesar_gastos(df_raw: pd.DataFrame):
    """
    Estructura esperada de gastos.xls (header=0):
      0 Fecha
      1 Docum.
      2 Cód. Cta. Ing.
      3 Nombre cta. ing.
      4 Cód. cuenta
      5 Título cta. perceptor
      6 CC
      7 CIF/NIF
      8 Base imponible
      9 I%
     10 Cuota
     11 Cuota rec. agr.
     12 Retención
     13 Observaciones
    """
    registros = []
    for _, row in df_raw.iterrows():
        # Filtro principal: debe tener CIF/NIF y base imponible numérica
        cif = limpiar_texto(row.iloc[7])
        base = row.iloc[8]
        if not cif or not es_numero_valido(base):
            continue

        fecha = formatear_fecha(row.iloc[0])
        vencimiento = calcular_vencimiento(fecha, dias=30)
        proveedor = limpiar_texto(row.iloc[5])
        nif = limpiar_nif(cif)
        factura = limpiar_texto(row.iloc[1])
        precio = parsear_precio(base)
        iva = parsear_iva(row.iloc[9])

        # Concepto de la línea: "Nombre cta. ing." según requerimiento
        concepto = limpiar_texto(row.iloc[3])
        if not concepto:
            observaciones = limpiar_texto(row.iloc[13])
            if observaciones:
                concepto = observaciones
            elif proveedor:
                concepto = f"Factura {factura} - {proveedor}"
            else:
                concepto = f"Factura {factura}"

        registros.append(
            {
                "Nombre Proveedor": proveedor,
                "NIF/CIF Proveedor": nif,
                "Numero Factura": factura,
                "Fecha Factura": fecha,
                "Fecha Vencimiento": vencimiento,
                "Concepto / Descripcion": concepto,
                "Cantidad": 1,
                "Precio Unitario": precio,
                "Cuenta Contable": "",
                "IVA (%)": iva,
                "Diario Compras (Codigo)": "Compras",
                "Empresa ID": 1,
            }
        )
    return pd.DataFrame(registros)


def procesar_ingresos(df_raw: pd.DataFrame):
    """
    Estructura esperada de ingresos.xls (header=1):
      0 Factura
      1 Fecha
      2 F.exped.
      3 Trim.
      4 Asiento
      5 Cód. cuenta
      6 Título cuenta
      7 CIF/NIF
      8 Base imponible
      9 I%
     10 Cuota IVA
     11 Retención
     12 Total
     13 Observaciones
     14 T. IVA
    """
    registros = []
    for _, row in df_raw.iterrows():
        # Filtro principal: debe tener CIF/NIF y base imponible numérica
        cif = limpiar_texto(row.iloc[7])
        base = row.iloc[8]
        if not cif or not es_numero_valido(base):
            continue

        fecha = formatear_fecha(row.iloc[1])
        vencimiento = calcular_vencimiento(fecha, dias=30)
        cliente = limpiar_texto(row.iloc[6])
        nif = limpiar_nif(cif)
        factura = limpiar_texto(row.iloc[0])
        precio = parsear_precio(base)
        iva = parsear_iva(row.iloc[9])

        observaciones = limpiar_texto(row.iloc[13])
        if not observaciones:
            if cliente:
                observaciones = f"Factura {factura} - {cliente}"
            else:
                observaciones = f"Factura {factura}"

        registros.append(
            {
                "Nombre Proveedor": cliente,
                "NIF/CIF Proveedor": nif,
                "Numero Factura": factura,
                "Fecha Factura": fecha,
                "Fecha Vencimiento": vencimiento,
                "Concepto / Descripcion": observaciones,
                "Cantidad": 1,
                "Precio Unitario": precio,
                "Cuenta Contable": "",
                "IVA (%)": iva,
                "Diario Compras (Codigo)": "Ventas",
                "Empresa ID": 1,
            }
        )
    return pd.DataFrame(registros)


# ---------------------------------------------------------------------------
# UI Streamlit
# ---------------------------------------------------------------------------
tipo = st.sidebar.selectbox("Tipo de archivo de entrada", ["Gasto", "Ingreso"])
uploaded_file = st.sidebar.file_uploader("Sube el archivo Excel (.xls o .xlsx)", type=["xls", "xlsx"])

if uploaded_file is not None:
    st.sidebar.success(f"Archivo cargado: {uploaded_file.name}")

    try:
        if tipo == "Gasto":
            df_raw = pd.read_excel(uploaded_file, header=0)
            df_plantilla = procesar_gastos(df_raw)
        else:
            df_raw = pd.read_excel(uploaded_file, header=1)
            df_plantilla = procesar_ingresos(df_raw)

        if df_plantilla.empty:
            st.warning("No se encontraron registros válidos después del filtrado. Revisa que el archivo tenga datos en las columnas CIF/NIF y Base imponible.")
        else:
            st.subheader(f"Vista previa de la plantilla ({tipo.lower()}s)")
            st.info(f"Total de registros convertidos: **{len(df_plantilla)}**")

            edited_df = st.data_editor(
                df_plantilla,
                num_rows="dynamic",
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Precio Unitario": st.column_config.NumberColumn(format="%.2f"),
                    "IVA (%)": st.column_config.NumberColumn(format="%d"),
                    "Cantidad": st.column_config.NumberColumn(format="%d"),
                    "Empresa ID": st.column_config.NumberColumn(format="%d"),
                },
            )

            st.markdown("---")
            excel_bytes = exportar_excel(edited_df)
            st.download_button(
                label="📥 Descargar plantilla_facturas.xlsx",
                data=excel_bytes,
                file_name="plantilla_facturas.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            st.markdown(
                "<small>Notas: El vencimiento se calcula como Fecha + 30 días. La columna 'Cuenta Contable' se deja vacía para rellenar manualmente antes de importar.</small>",
                unsafe_allow_html=True,
            )
    except Exception as e:
        st.error(f"Error al procesar el archivo: {e}")
        st.exception(e)
else:
    st.info("Selecciona el tipo de archivo y súbelo desde el panel lateral para comenzar.")
