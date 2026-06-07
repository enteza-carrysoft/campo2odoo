#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Convierte gastos.xls (listado de facturas de contabilidad) al formato
plantilla_facturas.xlsx requerido por Campo2Odoo.

Uso:
    python convertir_gastos_a_plantilla.py

El script lee 'gastos.xls' del directorio actual y genera
'plantilla_facturas.xlsx' listo para importar.
"""

import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re
import os

INPUT_FILE = "gastos.xls"
OUTPUT_FILE = "plantilla_facturas.xlsx"
DEFAULT_COMPANY_ID = 1
DEFAULT_JOURNAL_CODE = "Compras"

# Indices de columnas en gastos.xls (0-based)
COL_FACTURA = 0
COL_FECHA = 1
COL_COD_CUENTA = 5
COL_TITULO_CUENTA = 6
COL_CIF = 7
COL_BASE = 8
COL_IVA_PCT = 9
COL_OBS = 13


def es_fila_factura(row):
    """Detecta si una fila es una factura real (no resumen/total)."""
    factura = str(row.iloc[COL_FACTURA]).strip()
    fecha = row.iloc[COL_FECHA]
    cif = str(row.iloc[COL_CIF]).strip()
    base = row.iloc[COL_BASE]

    if not factura or not cif:
        return False
    try:
        float(base)
    except (ValueError, TypeError):
        return False

    if not isinstance(fecha, datetime) and not isinstance(fecha, pd.Timestamp):
        try:
            pd.to_datetime(fecha)
        except Exception:
            return False

    if re.search(r'imponible|total|recargo|resumen', factura, re.IGNORECASE):
        return False

    if isinstance(fecha, (int, float)):
        return False

    return True


def limpiar_texto(valor):
    if pd.isna(valor):
        return ""
    return str(valor).strip()


def limpiar_nif(nif):
    nif = limpiar_texto(nif)
    nif = re.sub(r'[\s\-.]', '', nif).upper()
    return nif


def formatear_fecha(valor):
    if pd.isna(valor):
        return ""
    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y")
    try:
        dt = pd.to_datetime(valor)
        return dt.strftime("%d/%m/%Y")
    except Exception:
        return str(valor).split()[0]


def calcular_vencimiento(fecha_str, dias=30):
    if not fecha_str:
        return ""
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(fecha_str, fmt)
            venc = dt + timedelta(days=dias)
            return venc.strftime("%d/%m/%Y")
        except ValueError:
            continue
    return ""


def deducir_descripcion(row):
    obs = limpiar_texto(row.iloc[COL_OBS])
    proveedor = limpiar_texto(row.iloc[COL_TITULO_CUENTA])
    factura = limpiar_texto(row.iloc[COL_FACTURA])
    if obs:
        return obs
    if proveedor:
        return f"Factura {factura} - {proveedor}"
    return f"Factura {factura}"


def deducir_cuenta_contable(cod_cuenta):
    cod = limpiar_texto(cod_cuenta)
    if cod.startswith("430"):
        return ""
    return cod


def convertir():
    if not os.path.exists(INPUT_FILE):
        print("[ERROR] No se encontro '{}' en el directorio actual.".format(INPUT_FILE))
        print("   Directorio actual: {}".format(os.getcwd()))
        return False

    print("[INFO] Leyendo '{}'...".format(INPUT_FILE))

    df = pd.read_excel(INPUT_FILE, header=1)
    print("   Total filas: {}".format(len(df)))

    facturas_df = df[df.apply(es_fila_factura, axis=1)].copy()
    print("   Filas de facturas validas: {}".format(len(facturas_df)))

    if len(facturas_df) == 0:
        print("[ADVERTENCIA] No se encontraron facturas validas.")
        return False

    registros = []
    for _, row in facturas_df.iterrows():
        fecha = formatear_fecha(row.iloc[COL_FECHA])
        vencimiento = calcular_vencimiento(fecha, dias=30)
        proveedor = limpiar_texto(row.iloc[COL_TITULO_CUENTA])
        nif = limpiar_nif(row.iloc[COL_CIF])
        factura = limpiar_texto(row.iloc[COL_FACTURA])
        base = row.iloc[COL_BASE]
        iva_pct = row.iloc[COL_IVA_PCT]
        observaciones = deducir_descripcion(row)
        cod_cuenta = deducir_cuenta_contable(row.iloc[COL_COD_CUENTA])

        try:
            precio = float(base) if not pd.isna(base) else 0.0
        except (ValueError, TypeError):
            precio = 0.0

        try:
            iva = int(float(iva_pct)) if not pd.isna(iva_pct) else 0
        except (ValueError, TypeError):
            iva = 0

        registros.append({
            "Nombre Proveedor": proveedor,
            "NIF/CIF Proveedor": nif,
            "Numero Factura": factura,
            "Fecha Factura": fecha,
            "Fecha Vencimiento": vencimiento,
            "Concepto / Descripcion": observaciones,
            "Cantidad": 1,
            "Precio Unitario": precio,
            "Cuenta Contable": cod_cuenta,
            "IVA (%)": iva,
            "Diario Compras (Codigo)": DEFAULT_JOURNAL_CODE,
            "Empresa ID": DEFAULT_COMPANY_ID,
        })

    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas"
    ws.sheet_view.showGridLines = True

    headers = list(registros[0].keys())
    ws.append(headers)

    for r in registros:
        ws.append(list(r.values()))

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    col_widths = {
        "A": 28, "B": 20, "C": 18, "D": 16, "E": 20,
        "F": 40, "G": 12, "H": 16, "I": 18, "J": 12,
        "K": 26, "L": 14,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
            if cell.column_letter == "G":
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="center")
            elif cell.column_letter == "H":
                cell.number_format = "#,##0.00"
            elif cell.column_letter == "J":
                cell.number_format = "0"
                cell.alignment = Alignment(horizontal="center")
            elif cell.column_letter == "L":
                cell.number_format = "0"
                cell.alignment = Alignment(horizontal="center")
            elif cell.column_letter in ("B", "C", "D", "E", "I", "K"):
                cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(OUTPUT_FILE)
    print("\n[OK] Archivo generado: '{}'".format(OUTPUT_FILE))
    print("   Facturas convertidas: {}".format(len(registros)))
    print("\n[NOTAS]")
    print("   - Cada factura genera UNA linea (cantidad=1, precio=base imponible).")
    print("   - La columna 'Cuenta Contable' se deja VACIA porque el archivo original")
    print("     contiene cuentas de proveedor (430...) en lugar de cuentas de gasto.")
    print("   - Revisa y rellena las cuentas contables (600..., 621..., 623...) antes")
    print("     de importar en Campo2Odoo.")
    print("   - El vencimiento se calcula como Fecha + 30 dias.")
    return True


if __name__ == "__main__":
    convertir()
